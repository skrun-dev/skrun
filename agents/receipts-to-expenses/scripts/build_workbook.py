#!/usr/bin/env python3
"""Render line items + category totals to expenses.xlsx (multi-sheet) + monthly.pdf (summary).

Input  (stdin JSON):  { line_items, category_totals, report_title, period, unmatched_count? }
Output (stdout JSON): { xlsx_path, pdf_path, total_amount, receipt_count }
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib import colors as rl_colors
except ImportError as e:
    print(json.dumps({
        "error": f"required deps not installed: {e}. Run: pip install -r requirements.txt"
    }))
    sys.exit(0)


BRAND = HexColor("#0366d6")
MUTED = HexColor("#666666")
HEADER_FILL = PatternFill(start_color="0366D6", end_color="0366D6", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_xlsx(out_path: Path, line_items: list, category_totals: list, period: str) -> None:
    """Two sheets - Line Items, Category Totals."""
    wb = Workbook()

    # Sheet 1: Line Items
    ws = wb.active
    ws.title = "Line Items"
    headers = ["Source File", "Vendor", "Date", "Amount", "Currency", "Category", "Note", "Matched Bank Row"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")

    for item in line_items:
        ws.append([
            item.get("source_file", ""),
            item.get("vendor", ""),
            item.get("date", ""),
            item.get("amount", 0),
            item.get("currency", ""),
            item.get("category", ""),
            item.get("note", ""),
            item.get("matched_bank_row", ""),
        ])

    # Auto-width approximation
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Sheet 2: Category Totals
    ws2 = wb.create_sheet(title="Category Totals")
    cat_headers = ["Category", "Total", "Count"]
    ws2.append(cat_headers)
    for col_idx, _ in enumerate(cat_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for cat in category_totals:
        ws2.append([
            cat.get("category", ""),
            cat.get("total", 0),
            cat.get("count", 0),
        ])

    # Period as a single cell hint at the bottom
    ws2.append([])
    ws2.append([f"Period: {period}"])

    for col_idx in range(1, len(cat_headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 24

    wb.save(str(out_path))


def write_pdf(
    out_path: Path,
    report_title: str,
    period: str,
    line_items: list,
    category_totals: list,
    unmatched_count: int,
) -> int:
    """1-2 page summary. Returns page count."""
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=letter,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        title=report_title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=BRAND, spaceAfter=8)
    subtitle_style = ParagraphStyle(
        "subtitle", parent=styles["Normal"], fontSize=12, textColor=MUTED, spaceAfter=24
    )
    section_h = ParagraphStyle(
        "section_h",
        parent=styles["Heading2"],
        textColor=BRAND,
        spaceBefore=18,
        spaceAfter=8,
    )
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=11, leading=16, spaceAfter=8)

    story: list = []
    story.append(Paragraph(report_title, title_style))
    story.append(Paragraph(period, subtitle_style))

    # Quick stats
    total = sum(it.get("amount", 0) for it in line_items)
    receipt_count = len(line_items)
    currency = line_items[0].get("currency", "USD") if line_items else "USD"
    stats_text = (
        f"<b>{receipt_count}</b> receipts processed. "
        f"Total: <b>{currency} {total:,.2f}</b>. "
    )
    if unmatched_count > 0:
        stats_text += (
            f"<b>{unmatched_count}</b> bank statement rows did not match any receipt — "
            f"see the workbook's Line Items sheet."
        )
    elif unmatched_count == 0 and any(it.get("matched_bank_row") for it in line_items):
        stats_text += "All receipts were reconciled against the bank statement."
    story.append(Paragraph(stats_text, body_style))

    # Category totals table
    if category_totals:
        story.append(Paragraph("Spend by category", section_h))
        rows = [["Category", "Total", "Count"]]
        for cat in sorted(category_totals, key=lambda c: -c.get("total", 0)):
            rows.append([
                cat.get("category", ""),
                f"{cat.get('total', 0):,.2f}",
                str(cat.get("count", 0)),
            ])
        rows.append(["Total", f"{total:,.2f}", str(receipt_count)])
        tbl = Table(rows, colWidths=[2.8 * inch, 1.5 * inch, 1.0 * inch])
        tbl.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#0366d6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#ffffff")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("LINEBELOW", (0, 0), (-1, -2), 0.5, rl_colors.lightgrey),
                ("BACKGROUND", (0, -1), (-1, -1), HexColor("#f6f8fa")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ])
        )
        story.append(tbl)

    if not line_items:
        story.append(Spacer(1, 12))
        story.append(Paragraph("<i>No receipts found in receipts_dir.</i>", body_style))

    doc.build(story)
    return 1  # this template fits on one page for typical inputs


def main() -> None:
    raw = sys.stdin.read()
    try:
        args = json.loads(raw)
    except json.JSONDecodeError as e:
        print(json.dumps({"error": f"invalid stdin JSON: {e}"}))
        return

    line_items = args.get("line_items", [])
    category_totals = args.get("category_totals", [])
    report_title = args.get("report_title", "Expense Report")
    period = args.get("period", "")
    unmatched_count = int(args.get("unmatched_count", 0))

    if not isinstance(line_items, list):
        line_items = []
    if not isinstance(category_totals, list):
        category_totals = []

    output_dir = os.environ.get("SKRUN_OUTPUT_DIR")
    if not output_dir:
        print(json.dumps({"error": "SKRUN_OUTPUT_DIR is not set"}))
        return

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    xlsx_path = Path(output_dir) / "expenses.xlsx"
    pdf_path = Path(output_dir) / "monthly.pdf"

    write_xlsx(xlsx_path, line_items, category_totals, period)
    write_pdf(pdf_path, report_title, period, line_items, category_totals, unmatched_count)

    total = sum(it.get("amount", 0) for it in line_items)

    print(json.dumps({
        "xlsx_path": str(xlsx_path),
        "pdf_path": str(pdf_path),
        "total_amount": total,
        "receipt_count": len(line_items),
    }))


if __name__ == "__main__":
    main()
